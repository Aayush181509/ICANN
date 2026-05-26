"""Generate synthetic financial/accounting data for the ICAN CA RAG training.

Produces:
    data/generated/pdf/   10 fictional PDFs (audit memo, policies, minutes, ...)
    data/generated/xlsx/  10 multi-sheet workbooks (TB, GL, FAR, AR aging, ...)
    data/generated/csv/   10 transaction-style CSVs (sales, purchases, JE, ...)
    data/generated/DATA_DICTIONARY.md

All data is **fully synthetic** — no real company or individual is referenced.

Teaching-friendly issues are intentionally seeded so participants can find them:
    * duplicate invoices
    * missing PAN / VAT numbers
    * round-number transactions
    * late payments
    * related-party indicators
    * policy exceptions
    * budget variances
    * inconsistent vendor names

Run:
    python src/data_generation.py
"""
from __future__ import annotations

import random
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib import colors

# ---------------------------------------------------------------------------
# Reproducibility
# ---------------------------------------------------------------------------
SEED = 20260526
random.seed(SEED)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data" / "generated"
PDF_DIR = DATA_DIR / "pdf"
XLSX_DIR = DATA_DIR / "xlsx"
CSV_DIR = DATA_DIR / "csv"
for d in (PDF_DIR, XLSX_DIR, CSV_DIR):
    d.mkdir(parents=True, exist_ok=True)


# ===========================================================================
# Fictional company universe
# ===========================================================================
COMPANY = "Himal Trading & Manufacturing Pvt. Ltd."
COMPANY_PAN = "300123456"
FY = "FY 2081/82"   # Nepali FY convention
REPORT_DATE = "2082-03-31"   # year-end

VENDORS = [
    # (vendor_id, name, PAN, related_party?)
    ("V001", "Kathmandu Steel Suppliers Pvt. Ltd.", "301234567", False),
    ("V002", "Pokhara Logistics Co.", "302345678", False),
    ("V003", "Everest Office Supplies", "303456789", False),
    ("V004", "Annapurna Holdings Pvt. Ltd.", "304567890", True),   # related party
    ("V005", "Himalayan IT Services", "305678901", False),
    ("V006", "Lumbini Packaging Industries", None, False),         # missing PAN
    ("V007", "Bagmati Power & Fuel", "307890123", False),
    ("V008", "Janakpur Spares & Tools", "308901234", False),
    ("V009", "Karnali Construction", "309012345", False),
    ("V010", "Himal Family Enterprises", "310123456", True),       # related party
    ("V011", "Kathmandu Steel Supplier P. Ltd.", "301234567", False),  # spelling variant of V001
    ("V012", "Mt. Everest Office Supply", "303456789", False),         # spelling variant of V003
]

CUSTOMERS = [
    ("C001", "Sagarmatha Distributors", "401234567"),
    ("C002", "Manaslu Retail Chain", "402345678"),
    ("C003", "Dhaulagiri Wholesale", "403456789"),
    ("C004", "Annapurna Holdings Pvt. Ltd.", "404567890"),  # related party customer
    ("C005", "Kanchanjunga Mart", "405678901"),
    ("C006", "Langtang Stores", None),                       # missing PAN
    ("C007", "Makalu Trading", "407890123"),
    ("C008", "Rara Exporters", "408901234"),
]

EMPLOYEES = [
    ("E001", "Ramesh Shrestha", "CEO"),
    ("E002", "Sita Karki", "CFO"),
    ("E003", "Bikram Thapa", "Procurement Manager"),
    ("E004", "Anju Pradhan", "Accounts Manager"),
    ("E005", "Deepak Adhikari", "Internal Auditor"),
    ("E006", "Meera Gurung", "Store Officer"),
    ("E007", "Prakash Rai", "HR Manager"),
    ("E008", "Sunita Maharjan", "Junior Accountant"),
]

ACCOUNTS = [
    ("1100", "Cash and Bank", "Asset"),
    ("1200", "Accounts Receivable", "Asset"),
    ("1300", "Inventory", "Asset"),
    ("1500", "Property, Plant and Equipment", "Asset"),
    ("1510", "Accumulated Depreciation", "Asset"),
    ("2100", "Accounts Payable", "Liability"),
    ("2200", "Loans Payable", "Liability"),
    ("2300", "VAT Payable", "Liability"),
    ("2400", "Income Tax Payable", "Liability"),
    ("3100", "Share Capital", "Equity"),
    ("3200", "Retained Earnings", "Equity"),
    ("4100", "Sales Revenue", "Income"),
    ("4200", "Other Income", "Income"),
    ("5100", "Cost of Goods Sold", "Expense"),
    ("5200", "Salaries and Wages", "Expense"),
    ("5300", "Rent Expense", "Expense"),
    ("5400", "Utilities", "Expense"),
    ("5500", "Depreciation Expense", "Expense"),
    ("5600", "Interest Expense", "Expense"),
    ("5700", "Other Operating Expenses", "Expense"),
]


def npr(amount: float) -> str:
    """Format NPR with Indian-style grouping for readability."""
    s = f"{amount:,.0f}"
    return f"NPR {s}"


# ===========================================================================
# PDF documents
# ===========================================================================
styles = getSampleStyleSheet()
H1 = ParagraphStyle("H1", parent=styles["Heading1"], spaceAfter=12)
H2 = ParagraphStyle("H2", parent=styles["Heading2"], spaceAfter=8)
BODY = ParagraphStyle("Body", parent=styles["BodyText"], leading=14, spaceAfter=6)
SMALL = ParagraphStyle("Small", parent=styles["BodyText"], fontSize=9, textColor=colors.grey)


def _write_pdf(filename: str, title: str, blocks: list):
    """Render a list of flowables to a PDF file."""
    path = PDF_DIR / filename
    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title=title,
        author=COMPANY,
    )
    story = [Paragraph(title, H1)]
    story.extend(blocks)
    story.append(Spacer(1, 12))
    story.append(Paragraph(
        f"<i>This is a synthetic teaching document. No real entity is referenced. {COMPANY} is fictional.</i>",
        SMALL,
    ))
    doc.build(story)


def _para(text: str):
    return Paragraph(text, BODY)


def _h2(text: str):
    return Paragraph(text, H2)


def _table(headers, rows, col_widths=None):
    data = [headers] + rows
    t = Table(data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    return t


def make_annual_report():
    blocks = [
        _para(f"<b>Company:</b> {COMPANY} (PAN: {COMPANY_PAN})<br/>"
              f"<b>Reporting Period:</b> {FY}, year ended {REPORT_DATE}<br/>"
              f"<b>Auditor:</b> XYZ &amp; Co., Chartered Accountants"),
        _h2("Director's Report (Extract)"),
        _para(
            "The Board is pleased to present the audited financial statements for "
            f"{FY}. Revenue grew 12.4% to NPR 48.6 crore on the back of strong "
            "domestic demand and the commissioning of the second production line. "
            "Profit after tax was NPR 4.2 crore (previous year: NPR 3.6 crore). "
            "The Board recommends a final dividend of NPR 5 per share."
        ),
        _h2("Key Financial Highlights"),
        _table(
            ["Indicator", FY, "FY 2080/81"],
            [
                ["Revenue", "486,234,512", "432,876,109"],
                ["Gross Profit", "121,558,628", "108,219,027"],
                ["EBITDA", "62,193,455", "55,442,098"],
                ["Profit After Tax", "42,114,872", "36,002,341"],
                ["Total Assets", "612,558,711", "548,219,440"],
                ["Total Borrowings", "182,000,000", "165,000,000"],
            ],
        ),
        _h2("Borrowings Disclosure"),
        _para(
            "The Company has outstanding term loans of NPR 18.2 crore from "
            "Nepal Commercial Bank Ltd. The loan is repayable over 7 years at "
            "an interest rate of 10.5% per annum. Refer to the Loan Agreement "
            "Summary (separate document) for the full covenant schedule. "
            "Cross-reference: Note 14 of the audited accounts; Loan Schedule "
            "register maintained by the Finance department."
        ),
        _h2("Related Party Transactions"),
        _para(
            "During the year the Company transacted with Annapurna Holdings "
            "Pvt. Ltd. (an entity controlled by the Chairman's family) and "
            "Himal Family Enterprises (an entity in which a director has "
            "significant influence). Refer to the Related Party Transaction "
            "Listing for the full break-up. Aggregate transactions: "
            "purchases NPR 1.86 crore, sales NPR 0.42 crore, year-end balances "
            "receivable NPR 28.5 lakh."
        ),
        _h2("Going Concern"),
        _para(
            "The Board has reviewed the cash-flow projections for the next 18 "
            "months and is satisfied that the Company will continue as a going "
            "concern. Loan covenant headroom is monitored on a quarterly basis."
        ),
    ]
    _write_pdf("01_annual_report_extract.pdf", f"Annual Report Extract — {COMPANY} — {FY}", blocks)


def make_audit_planning_memo():
    blocks = [
        _para(f"<b>To:</b> Engagement Team<br/>"
              f"<b>From:</b> Engagement Partner, XYZ &amp; Co.<br/>"
              f"<b>Re:</b> Audit Planning Memorandum — {COMPANY} — {FY}<br/>"
              f"<b>Date:</b> 2082-04-10"),
        _h2("1. Engagement scope"),
        _para(
            "We have been re-appointed as statutory auditors for the financial "
            f"year ended {REPORT_DATE}. The audit will be conducted in accordance "
            "with Nepal Standards on Auditing (NSAs). Reporting deadline: "
            "2082-06-15."
        ),
        _h2("2. Preliminary risk assessment"),
        _para(
            "Based on prior-year working papers, analytical review and discussion "
            "with management, the following are considered <b>Significant Risks</b>:"
        ),
        _table(
            ["#", "Risk area", "Why significant", "Planned response"],
            [
                ["R1", "Revenue recognition — cut-off",
                 "Q4 revenue spike of 28% vs run-rate",
                 "Cut-off testing on 15 days either side of YE"],
                ["R2", "Inventory existence and valuation",
                 "Second production line commissioned; high WIP balances",
                 "Attend physical count; NRV testing on slow-moving SKUs"],
                ["R3", "Related party transactions",
                 "Family-controlled entities; round-amount transactions noted",
                 "Inspect approvals; confirm balances; review minutes"],
                ["R4", "Loan covenant compliance",
                 "DSCR covenant of 1.25x; tight headroom",
                 "Re-compute DSCR; obtain bank confirmation"],
                ["R5", "Going concern",
                 "Working capital pressure observed at interim",
                 "Review cash-flow forecast; sensitivity testing"],
            ],
            col_widths=[1*cm, 4*cm, 5.5*cm, 5.5*cm],
        ),
        _h2("3. Materiality"),
        _para(
            "Overall materiality has been set at NPR 21,00,000 (approximately "
            "5% of profit before tax). Performance materiality has been set at "
            "75% of overall materiality, i.e. NPR 15,75,000. Clearly trivial "
            "threshold is NPR 1,05,000."
        ),
        _h2("4. Team composition"),
        _para(
            "Engagement partner — CA Ramesh K.C.; Manager — CA Sushma Lama; "
            "In-charge — CA Article Niraj Karki; Trainees — 2."
        ),
        _h2("5. Cross-references"),
        _para(
            "See: Internal Control Policy (separate document) for control "
            "design; Loan Agreement Summary for covenant clauses; Board Minutes "
            "for related-party approvals."
        ),
    ]
    _write_pdf("02_audit_planning_memo.pdf", f"Audit Planning Memorandum — {COMPANY}", blocks)


def make_internal_control_policy():
    blocks = [
        _para(f"<b>Document:</b> Internal Control Policy<br/>"
              f"<b>Owner:</b> Chief Financial Officer<br/>"
              f"<b>Effective date:</b> 2081-04-01<br/>"
              f"<b>Next review:</b> 2082-04-01"),
        _h2("1. Purpose"),
        _para(
            "This policy sets out the minimum internal control standards to "
            f"be observed by all departments of {COMPANY}. It is aligned with "
            "the COSO 2013 framework and ICAN's guidance on internal control."
        ),
        _h2("2. Segregation of duties"),
        _para(
            "No single individual shall (a) initiate, (b) approve, and (c) "
            "record any financial transaction. Reviewer and approver must be "
            "different persons. Where this is impractical (small department), "
            "compensating controls — independent review or quarterly rotation — "
            "must be documented."
        ),
        _h2("3. Approval matrix"),
        _table(
            ["Transaction type", "Threshold (NPR)", "Approver"],
            [
                ["Purchase requisition", "Up to 1,00,000", "Department Head"],
                ["Purchase requisition", "1,00,001 — 5,00,000", "Procurement Manager + CFO"],
                ["Purchase requisition", "Above 5,00,000", "CEO + Board sub-committee"],
                ["Journal entry (non-routine)", "Any", "Accounts Manager"],
                ["Payment", "Above 10,00,000", "CFO + one Director"],
                ["Related party transaction", "Any", "Board approval, declared in minutes"],
                ["Write-off / waiver", "Any", "CFO + Board ratification"],
            ],
        ),
        _h2("4. Vendor master controls"),
        _para(
            "All new vendors must be on-boarded through the Procurement "
            "department with KYC documents: PAN/VAT registration certificate, "
            "bank details, registered office address. The Accounts Payable "
            "team shall not process invoices from vendors not on the approved "
            "list. Duplicate vendor codes are prohibited."
        ),
        _h2("5. Cash and bank"),
        _para(
            "Bank reconciliation must be performed monthly within 5 working "
            "days of month-end and signed off by the Accounts Manager. "
            "Cash holdings at any business location must not exceed NPR 50,000."
        ),
        _h2("6. Policy exceptions"),
        _para(
            "Any deviation from this policy requires written approval from "
            "the CFO and must be logged in the Exception Register. The CFO "
            "shall present a summary of exceptions to the Audit Committee "
            "on a quarterly basis."
        ),
    ]
    _write_pdf("03_internal_control_policy.pdf", "Internal Control Policy", blocks)


def make_procurement_policy():
    blocks = [
        _para(f"<b>Document:</b> Procurement Policy<br/>"
              f"<b>Owner:</b> Procurement Manager<br/>"
              f"<b>Effective date:</b> 2081-07-01"),
        _h2("1. Vendor selection"),
        _para(
            "For purchases above NPR 1,00,000 a minimum of three competitive "
            "quotations must be obtained. For purchases above NPR 10,00,000 a "
            "sealed-bid tender process is required. Sole-source procurement "
            "is permitted only with the CFO's prior written approval and "
            "documented justification."
        ),
        _h2("2. Three-way match"),
        _para(
            "Accounts Payable shall not process an invoice for payment unless "
            "(a) a Purchase Order, (b) a Goods Receipt Note, and (c) the "
            "Vendor Invoice are matched on quantity, rate and total amount. "
            "Tolerances: 2% on rate, zero on quantity."
        ),
        _h2("3. Related party purchases"),
        _para(
            "Any procurement from a related party (as defined in the Internal "
            "Control Policy) requires <b>Board approval before order placement</b> "
            "and disclosure in the Related Party Transaction Listing. Three-quote "
            "exemption does not apply; an independent benchmark price must be "
            "obtained and retained on file."
        ),
        _h2("4. Vendor master"),
        _para(
            "PAN/VAT registration is mandatory for vendor master creation. "
            "Procurement shall maintain copies of registration certificates "
            "and conduct an annual re-validation."
        ),
        _h2("5. Exceptions"),
        _para(
            "Emergency procurements may be made without prior approval up to "
            "NPR 50,000, subject to ratification within 7 working days."
        ),
    ]
    _write_pdf("04_procurement_policy.pdf", "Procurement Policy", blocks)


def make_tax_compliance_memo():
    blocks = [
        _para(f"<b>To:</b> CFO<br/><b>From:</b> Tax Advisor<br/>"
              f"<b>Re:</b> Tax compliance status — {FY}<br/>"
              f"<b>Date:</b> 2082-04-20"),
        _h2("1. Income tax"),
        _para(
            "Provisional self-assessment for income tax has been computed at "
            "NPR 1.81 crore. The Company has paid advance tax of NPR 1.65 crore. "
            "Balance tax of NPR 16 lakh is payable by 2082-06-30 along with the "
            "annual return."
        ),
        _h2("2. VAT"),
        _para(
            "Monthly VAT returns are up to date. Input credit reconciliation "
            "between the GL VAT-Receivable account and the IRD portal is "
            "current except for Ashad (NPR 2.3 lakh under review)."
        ),
        _h2("3. TDS"),
        _para(
            "TDS deductions on rent, contractor payments and professional fees "
            "are reconciled monthly. Two cases of late deposit (Mangsir and "
            "Magh) attracted penalty of NPR 14,500 in total, which has been "
            "paid."
        ),
        _h2("4. Open exposures"),
        _table(
            ["Issue", "Estimated exposure (NPR)", "Status"],
            [
                ["VAT input credit on capital goods — IRD query", "8,50,000", "Reply filed; awaiting hearing"],
                ["Transfer pricing on related party purchases", "Indeterminate", "Documentation to be prepared"],
                ["Income tax assessment — FY 2079/80", "12,00,000", "Appeal pending"],
            ],
        ),
        _h2("5. Recommendations"),
        _para(
            "(a) Document the transfer-pricing basis for all related party "
            "transactions before 2082-06-15. (b) Set up an automated TDS "
            "deposit reminder. (c) Consider obtaining an advance ruling on "
            "the capital-goods input credit question."
        ),
    ]
    _write_pdf("05_tax_compliance_memo.pdf", f"Tax Compliance Memo — {FY}", blocks)


def make_loan_agreement_summary():
    blocks = [
        _para(f"<b>Borrower:</b> {COMPANY}<br/>"
              f"<b>Lender:</b> Nepal Commercial Bank Ltd.<br/>"
              f"<b>Facility:</b> Term Loan<br/>"
              f"<b>Sanction date:</b> 2080-08-15<br/>"
              f"<b>Sanctioned amount:</b> NPR 20,00,00,000"),
        _h2("Key commercial terms"),
        _table(
            ["Term", "Value"],
            [
                ["Outstanding (at year-end)", "NPR 18,20,00,000"],
                ["Tenor", "7 years (84 months)"],
                ["Interest rate", "10.5% p.a., reset annually to base + 3.25%"],
                ["Moratorium", "12 months (already lapsed)"],
                ["Repayment", "84 EMIs of approximately NPR 33,80,000"],
                ["Security", "Hypothecation of fixed assets; personal guarantee of two directors"],
            ],
        ),
        _h2("Financial covenants"),
        _para(
            "The borrower shall maintain, tested annually on the audited "
            "financial statements:<br/>"
            "<b>(a)</b> Debt-Service Coverage Ratio (DSCR) of not less than "
            "<b>1.25 times</b>.<br/>"
            "<b>(b)</b> Debt-to-Equity ratio not exceeding <b>2.0 times</b>.<br/>"
            "<b>(c)</b> Current ratio not less than <b>1.10 times</b>."
        ),
        _h2("Affirmative covenants"),
        _para(
            "The borrower shall furnish: (i) annual audited financials within "
            "6 months of year-end; (ii) quarterly management accounts within "
            "45 days; (iii) prior written intimation of any related-party "
            "transaction exceeding NPR 1,00,00,000."
        ),
        _h2("Negative covenants"),
        _para(
            "Without the prior written consent of the lender, the borrower "
            "shall not: (i) declare dividend if DSCR &lt; 1.50x; (ii) incur "
            "any additional secured borrowing; (iii) change shareholding "
            "structure by more than 25%."
        ),
        _h2("Events of default"),
        _para(
            "Any breach of covenant, payment default exceeding 30 days, or "
            "change of management without intimation constitutes an event of "
            "default."
        ),
    ]
    _write_pdf("06_loan_agreement_summary.pdf", "Loan Agreement Summary", blocks)


def make_board_minutes():
    blocks = [
        _para(f"<b>Meeting:</b> 142nd Meeting of the Board of Directors<br/>"
              f"<b>Date:</b> 2082-02-18<br/>"
              f"<b>Venue:</b> Registered Office, Kathmandu<br/>"
              f"<b>Chair:</b> Mr. Ramesh Shrestha"),
        _h2("Present"),
        _para(
            "Mr. Ramesh Shrestha (Chair, Executive Director), Ms. Sita Karki "
            "(Executive Director — Finance), Mr. Krishna Bhattarai (Independent "
            "Director), Ms. Asha Subba (Independent Director), Mr. Bishal Joshi "
            "(Non-Executive Director). In attendance: Internal Auditor."
        ),
        _h2("Item 1 — Confirmation of previous minutes"),
        _para("The minutes of the 141st meeting were confirmed without amendment."),
        _h2("Item 2 — Q3 financial performance"),
        _para(
            "The CFO presented Q3 results. Revenue is tracking 11% above prior "
            "year. EBITDA margin is broadly stable. Trade receivables ageing "
            "shows a moderate increase in the &gt;90-day bucket, attributed to "
            "two customers; provision policy will be reviewed at year-end."
        ),
        _h2("Item 3 — Related party transactions"),
        _para(
            "The Board reviewed and approved the following related party "
            "transactions for the year:<br/>"
            "(a) Purchases from <b>Annapurna Holdings Pvt. Ltd.</b> up to "
            "NPR 1,50,00,000 — approved.<br/>"
            "(b) Purchases from <b>Himal Family Enterprises</b> up to "
            "NPR 50,00,000 — approved subject to obtaining benchmark prices.<br/>"
            "(c) Sales to <b>Annapurna Holdings Pvt. Ltd.</b> up to NPR 50,00,000 "
            "— approved."
        ),
        _h2("Item 4 — Loan covenant compliance"),
        _para(
            "The CFO confirmed that DSCR is currently estimated at 1.32x, "
            "above the 1.25x covenant, but with limited headroom. The Board "
            "advised conservative dividend policy this year."
        ),
        _h2("Item 5 — Internal audit findings"),
        _para(
            "The Internal Auditor presented findings from the procurement "
            "review. Two cases of sole-source procurement above policy "
            "threshold were noted; the CFO was requested to investigate and "
            "report at the next meeting."
        ),
        _h2("Item 6 — Any other business"),
        _para(
            "None. Meeting concluded at 14:30. Next meeting: 2082-05-20."
        ),
    ]
    _write_pdf("07_board_minutes.pdf", "Minutes of the 142nd Board Meeting", blocks)


def make_management_letter():
    blocks = [
        _para(f"<b>Issued by:</b> XYZ &amp; Co., Chartered Accountants<br/>"
              f"<b>To:</b> The Board of Directors, {COMPANY}<br/>"
              f"<b>Re:</b> Management Letter — {FY}<br/>"
              f"<b>Date:</b> 2082-06-10"),
        _para(
            "In the course of our audit, we noted the following matters which "
            "we are required to communicate to you. These do not affect our "
            "audit opinion but are brought to your attention for corrective "
            "action."
        ),
        _h2("1. Vendor master — duplicate entries"),
        _para(
            "We identified duplicate vendor codes for the same vendor "
            "(<i>Kathmandu Steel Suppliers</i> and <i>Mt. Everest Office "
            "Supply</i>) which could lead to duplicate payments. We recommend "
            "an immediate vendor master clean-up and a control to prevent "
            "creation of duplicate vendors at on-boarding stage."
        ),
        _h2("2. Approval threshold breaches"),
        _para(
            "Three purchase transactions exceeding NPR 5,00,000 lacked the "
            "required dual approval per the Internal Control Policy. Refer to "
            "Annexure A for details."
        ),
        _h2("3. Related party documentation"),
        _para(
            "Benchmark pricing documentation for purchases from <i>Himal Family "
            "Enterprises</i> was incomplete. This is contrary to the Procurement "
            "Policy and creates a transfer-pricing exposure."
        ),
        _h2("4. PAN/VAT registration"),
        _para(
            "Two active vendors do not have PAN/VAT details captured in the "
            "vendor master. We recommend a one-time data validation exercise."
        ),
        _h2("5. Bank reconciliations"),
        _para(
            "Bank reconciliations for the period Magh-Chaitra contained "
            "long-outstanding items (over 6 months). We recommend a monthly "
            "review of stale reconciling items by the CFO."
        ),
        _h2("6. IT general controls"),
        _para(
            "We observed shared user IDs being used in the accounting system. "
            "This compromises audit trail. Individual user IDs with role-based "
            "access should be implemented."
        ),
    ]
    _write_pdf("08_management_letter.pdf", f"Management Letter — {FY}", blocks)


def make_nfrs_policy_note():
    blocks = [
        _para(f"<b>Document:</b> Significant Accounting Policies (Extract)<br/>"
              f"<b>Framework:</b> Nepal Financial Reporting Standards (NFRS)"),
        _h2("1. Revenue recognition"),
        _para(
            "Revenue is recognised in accordance with NFRS 15 — <i>Revenue "
            "from Contracts with Customers</i>. Revenue is recognised when "
            "control of the goods is transferred to the customer, generally "
            "on dispatch from the factory gate (Incoterms: EXW). Sales "
            "returns are estimated based on historical experience and "
            "recognised as a reduction in revenue."
        ),
        _h2("2. Inventories"),
        _para(
            "Inventories are stated at the lower of cost and net realisable "
            "value (NFRS 2). Cost is determined on a weighted-average basis "
            "for raw materials and a standard-cost basis for finished goods "
            "(with periodic variance adjustments)."
        ),
        _h2("3. Property, Plant and Equipment"),
        _para(
            "PPE is stated at cost less accumulated depreciation. "
            "Depreciation is provided on a straight-line basis over the "
            "estimated useful lives: Buildings 30 years; Plant &amp; Machinery "
            "10 years; Vehicles 5 years; Furniture 7 years; Computer "
            "equipment 4 years."
        ),
        _h2("4. Financial instruments"),
        _para(
            "Financial assets and liabilities are recognised and measured in "
            "accordance with NFRS 9. Trade receivables are measured at "
            "amortised cost. Expected credit losses are estimated using a "
            "simplified provision matrix based on ageing buckets."
        ),
        _h2("5. Borrowing costs"),
        _para(
            "Borrowing costs directly attributable to the acquisition or "
            "construction of qualifying assets are capitalised. All other "
            "borrowing costs are expensed in the period in which they are "
            "incurred."
        ),
        _h2("6. Related party transactions"),
        _para(
            "Related party transactions are disclosed in accordance with "
            "NAS 24. All such transactions are conducted at arm's length "
            "where practicable. Where arm's-length pricing cannot be "
            "demonstrated, this fact is disclosed."
        ),
    ]
    _write_pdf("09_nfrs_accounting_policy_note.pdf", "Significant Accounting Policies — NFRS", blocks)


def make_inventory_observation_memo():
    blocks = [
        _para(f"<b>To:</b> Engagement file<br/>"
              f"<b>From:</b> In-charge, XYZ &amp; Co.<br/>"
              f"<b>Re:</b> Physical inventory observation — {REPORT_DATE}"),
        _h2("1. Locations and timing"),
        _para(
            "Physical inventory count was conducted on 2082-03-31 at the "
            "factory premises (Bhaktapur), central warehouse (Balaju) and "
            "regional depot (Pokhara). The audit team attended all three "
            "locations."
        ),
        _h2("2. Count procedures observed"),
        _para(
            "The management count team comprised store officers and finance "
            "representatives. Pre-numbered count sheets were used. Test counts "
            "on a sample basis (50 SKUs per location, selected on a stratified "
            "random basis) showed reasonable accuracy, with the exceptions "
            "noted below."
        ),
        _h2("3. Exceptions noted"),
        _table(
            ["Location", "SKU", "Book qty", "Counted qty", "Variance", "Indicative value (NPR)"],
            [
                ["Bhaktapur", "RM-014 Steel Rod 12mm", "1,250", "1,189", "(61) units", "(1,52,500)"],
                ["Balaju", "FG-302 Pipe 4-inch", "642", "658", "+16 units", "+88,000"],
                ["Pokhara", "FG-118 Sheet 1mm", "98", "84", "(14) units", "(56,000)"],
            ],
            col_widths=[2.5*cm, 4.0*cm, 2.2*cm, 2.5*cm, 2.5*cm, 3.5*cm],
        ),
        _h2("4. Slow-moving and obsolete"),
        _para(
            "Items aged more than 365 days amount to NPR 38.4 lakh (4.7% of "
            "total inventory). Management is currently providing 25% against "
            "this category; NFRS 2 (NRV testing) suggests this may be "
            "insufficient — recommend management revisit the provisioning "
            "policy."
        ),
        _h2("5. Cut-off"),
        _para(
            "Last GRN (2082-03-31): GRN/2082/0184. First GRN of new year "
            "(2082-04-01): GRN/2082/0185. Cut-off appears clean; no goods in "
            "transit at year-end."
        ),
    ]
    _write_pdf("10_inventory_observation_memo.pdf", "Inventory Observation Memo", blocks)


def generate_all_pdfs():
    make_annual_report()
    make_audit_planning_memo()
    make_internal_control_policy()
    make_procurement_policy()
    make_tax_compliance_memo()
    make_loan_agreement_summary()
    make_board_minutes()
    make_management_letter()
    make_nfrs_policy_note()
    make_inventory_observation_memo()


# ===========================================================================
# Excel (XLSX) workbooks
# ===========================================================================
def _save_workbook(wb: Workbook, name: str):
    path = XLSX_DIR / name
    wb.save(path)


def make_trial_balance():
    wb = Workbook()
    ws = wb.active
    ws.title = "TB"
    ws.append(["Account Code", "Account Name", "Type", "Debit (NPR)", "Credit (NPR)"])
    # Deterministic balances that *almost* balance — small intentional imbalance
    # for the cleaning exercise in Notebook 03.
    rows = [
        ("1100", "Cash and Bank", "Asset", 32_450_000, 0),
        ("1200", "Accounts Receivable", "Asset", 81_320_000, 0),
        ("1300", "Inventory", "Asset", 86_400_000, 0),
        ("1500", "Property, Plant and Equipment", "Asset", 312_000_000, 0),
        ("1510", "Accumulated Depreciation", "Asset", 0, 96_500_000),
        ("2100", "Accounts Payable", "Liability", 0, 42_800_000),
        ("2200", "Loans Payable", "Liability", 0, 182_000_000),
        ("2300", "VAT Payable", "Liability", 0, 5_200_000),
        ("2400", "Income Tax Payable", "Liability", 0, 16_000_000),
        ("3100", "Share Capital", "Equity", 0, 100_000_000),
        ("3200", "Retained Earnings", "Equity", 0, 27_350_000),
        ("4100", "Sales Revenue", "Income", 0, 486_234_500),
        ("4200", "Other Income", "Income", 0, 1_820_000),
        ("5100", "Cost of Goods Sold", "Expense", 364_676_000, 0),
        ("5200", "Salaries and Wages", "Expense", 38_500_000, 0),
        ("5300", "Rent Expense", "Expense", 7_200_000, 0),
        ("5400", "Utilities", "Expense", 4_600_000, 0),
        ("5500", "Depreciation Expense", "Expense", 18_900_000, 0),
        ("5600", "Interest Expense", "Expense", 19_110_000, 0),
        ("5700", "Other Operating Expenses", "Expense", 10_758_500, 0),
    ]
    for r in rows:
        ws.append(r)
    # Totals row (intentionally leave imbalance ~ 0 to ~ small)
    total_dr = sum(r[3] for r in rows)
    total_cr = sum(r[4] for r in rows)
    ws.append(["", "TOTAL", "", total_dr, total_cr])
    _save_workbook(wb, "01_trial_balance.xlsx")


def make_general_ledger():
    wb = Workbook()
    ws = wb.active
    ws.title = "GL_Sales"
    ws.append(["Date", "Voucher", "Account", "Description", "Debit", "Credit"])
    start = date(2081, 4, 1)
    for i in range(60):
        d = start + timedelta(days=random.randint(0, 360))
        amt = random.randint(50_000, 1_500_000)
        ws.append([d.isoformat(), f"SI-{1000+i}", "4100", "Sales invoice", 0, amt])
        ws.append([d.isoformat(), f"SI-{1000+i}", "1200", "AR", amt, 0])

    ws2 = wb.create_sheet("GL_Purchases")
    ws2.append(["Date", "Voucher", "Account", "Description", "Debit", "Credit"])
    for i in range(50):
        d = start + timedelta(days=random.randint(0, 360))
        amt = random.randint(30_000, 1_200_000)
        ws2.append([d.isoformat(), f"PI-{2000+i}", "5100", "Purchase invoice", amt, 0])
        ws2.append([d.isoformat(), f"PI-{2000+i}", "2100", "AP", 0, amt])

    ws3 = wb.create_sheet("GL_Adjustments")
    ws3.append(["Date", "Voucher", "Account", "Description", "Debit", "Credit", "Posted by"])
    adj = [
        ("2082-03-30", "JE-0091", "5100", "Reclass purchase to COGS",        500_000, 0, "E008"),
        ("2082-03-30", "JE-0091", "1300", "Inventory adj",                   0, 500_000, "E008"),
        ("2082-03-31", "JE-0095", "4100", "Reversal: invoice cancelled",     250_000, 0, "E004"),  # late, round
        ("2082-03-31", "JE-0095", "1200", "AR reversal",                     0, 250_000, "E004"),
        # Suspicious: round number JE posted late at night by junior, no description
        ("2082-03-31", "JE-0099", "5700", "Adjustment",                      1_000_000, 0, "E008"),
        ("2082-03-31", "JE-0099", "1100", "Bank",                            0, 1_000_000, "E008"),
    ]
    for row in adj:
        ws3.append(row)
    _save_workbook(wb, "02_general_ledger_sample.xlsx")


def make_fixed_asset_register():
    wb = Workbook()
    ws = wb.active
    ws.title = "FAR"
    ws.append(["Asset Code", "Description", "Category", "Acquisition Date",
               "Cost (NPR)", "Useful Life (yrs)", "Accumulated Dep (NPR)", "Net Book Value (NPR)"])
    rows = [
        ("FA-001", "Factory Building", "Building", "2076-04-15", 120_000_000, 30, 24_000_000, 96_000_000),
        ("FA-002", "Production Line 1", "Plant & Machinery", "2077-09-12", 78_000_000, 10, 31_200_000, 46_800_000),
        ("FA-003", "Production Line 2", "Plant & Machinery", "2081-08-01", 65_000_000, 10, 5_416_667, 59_583_333),
        ("FA-004", "Toyota Hilux", "Vehicle", "2079-02-10", 5_400_000, 5, 3_240_000, 2_160_000),
        ("FA-005", "Generator 250 kVA", "Plant & Machinery", "2080-11-22", 12_000_000, 10, 1_600_000, 10_400_000),
        ("FA-006", "Server room equipment", "Computer", "2080-06-30", 4_200_000, 4, 1_575_000, 2_625_000),
        ("FA-007", "Office Furniture (Lot)", "Furniture", "2078-04-01", 1_800_000, 7, 1_028_571, 771_429),
        ("FA-008", "Forklift", "Plant & Machinery", "2079-09-15", 3_800_000, 10, 950_000, 2_850_000),
    ]
    for r in rows:
        ws.append(r)
    _save_workbook(wb, "03_fixed_asset_register.xlsx")


def make_ar_aging():
    wb = Workbook()
    ws = wb.active
    ws.title = "AR_Aging"
    ws.append(["Customer Code", "Customer Name", "Current (NPR)",
               "1-30 days", "31-60 days", "61-90 days", ">90 days", "Total"])
    rows = [
        ("C001", "Sagarmatha Distributors",  6_200_000, 3_100_000, 1_200_000,   400_000,   200_000, 11_100_000),
        ("C002", "Manaslu Retail Chain",     4_800_000, 2_400_000,   600_000,   100_000,         0,  7_900_000),
        ("C003", "Dhaulagiri Wholesale",     3_500_000, 1_800_000,   900_000,   500_000,   320_000,  7_020_000),
        ("C004", "Annapurna Holdings",       1_200_000, 1_200_000, 1_200_000, 1_200_000, 1_200_000,  6_000_000),  # related party — flat aging
        ("C005", "Kanchanjunga Mart",        2_300_000,   600_000,         0,         0,         0,  2_900_000),
        ("C006", "Langtang Stores",          1_500_000,   300_000,   180_000,    50_000,         0,  2_030_000),
        ("C007", "Makalu Trading",           4_100_000,   800_000,   250_000,         0,         0,  5_150_000),
        ("C008", "Rara Exporters",                   0,         0, 1_500_000, 1_500_000, 2_700_000,  5_700_000),  # all overdue
    ]
    for r in rows:
        ws.append(r)
    _save_workbook(wb, "04_accounts_receivable_aging.xlsx")


def make_inventory_listing():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(["SKU", "Description", "Category", "Location", "Qty",
               "Unit Cost (NPR)", "Total Cost (NPR)", "Ageing (days)", "Last Movement"])
    items = [
        ("RM-001", "Steel Rod 8mm", "Raw Material", "Bhaktapur", 2400, 220, 528_000, 45, "2082-02-15"),
        ("RM-014", "Steel Rod 12mm", "Raw Material", "Bhaktapur", 1189, 480, 570_720, 30, "2082-03-01"),
        ("RM-027", "Coil 0.5mm", "Raw Material", "Bhaktapur", 540, 1_200, 648_000, 380, "2081-03-12"),  # slow moving
        ("WIP-101", "Pipe 4-inch in process", "WIP", "Bhaktapur", 320, 900, 288_000, 25, "2082-03-06"),
        ("FG-302", "Pipe 4-inch", "Finished Goods", "Balaju", 658, 1_600, 1_052_800, 18, "2082-03-13"),
        ("FG-118", "Sheet 1mm", "Finished Goods", "Pokhara", 84, 4_000, 336_000, 12, "2082-03-19"),
        ("FG-205", "Channel 50mm", "Finished Goods", "Balaju", 110, 3_200, 352_000, 410, "2081-02-20"),   # slow
        ("CONS-09", "Welding electrodes box", "Consumable", "Bhaktapur", 220, 850, 187_000, 60, "2082-01-30"),
        ("FG-509", "Custom flange (specials)", "Finished Goods", "Balaju", 18, 12_000, 216_000, 700, "2080-04-25"),  # obsolete
    ]
    for r in items:
        ws.append(r)
    _save_workbook(wb, "05_inventory_listing.xlsx")


def make_bank_reconciliation():
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank_Recon"
    ws.append(["Particulars", "Amount (NPR)"])
    ws.append(["Balance as per Bank statement (2082-03-31)", 31_840_500])
    ws.append(["Add: Cheques deposited not yet credited", 850_000])
    ws.append(["Less: Cheques issued not yet presented", (1_240_000)])
    ws.append(["Less: Bank charges in statement, not in books", (5_500)])
    ws.append(["Add: Direct deposit by customer not in books", 5_000])
    ws.append(["Balance as per Books (2082-03-31)", 31_450_000])

    ws2 = wb.create_sheet("Outstanding_Items")
    ws2.append(["Type", "Reference", "Date", "Amount (NPR)", "Age (days)", "Status"])
    items = [
        ("Cheque issued — not presented", "CHQ-19102", "2082-03-28", 540_000, 3, "Normal"),
        ("Cheque issued — not presented", "CHQ-18904", "2081-10-12", 200_000, 170, "Stale — investigate"),
        ("Cheque issued — not presented", "CHQ-17500", "2081-06-30", 500_000, 274, "Stale — investigate"),
        ("Cheque deposited — not credited", "CHQ-IN-7782", "2082-03-30", 850_000, 1, "Normal"),
    ]
    for r in items:
        ws2.append(r)
    _save_workbook(wb, "06_bank_reconciliation.xlsx")


def make_loan_schedule():
    wb = Workbook()
    ws = wb.active
    ws.title = "Loan_Schedule"
    ws.append(["Installment #", "Due Date", "Opening (NPR)", "EMI (NPR)",
               "Interest (NPR)", "Principal (NPR)", "Closing (NPR)"])
    principal = 200_000_000
    rate_m = 0.105 / 12
    emi = 3_380_000
    bal = principal
    d = date(2080, 9, 15)
    for i in range(1, 25):  # 24 months shown
        interest = round(bal * rate_m)
        prin = emi - interest
        new_bal = bal - prin
        ws.append([i, d.isoformat(), bal, emi, interest, prin, new_bal])
        bal = new_bal
        # next month (approximation)
        d = date(d.year + (1 if d.month == 12 else 0), 1 if d.month == 12 else d.month + 1, 15)

    ws2 = wb.create_sheet("Covenants")
    ws2.append(["Covenant", "Threshold", "Current", "Status"])
    ws2.append(["DSCR", ">= 1.25x", 1.32, "Compliant"])
    ws2.append(["Debt/Equity", "<= 2.00x", 1.43, "Compliant"])
    ws2.append(["Current Ratio", ">= 1.10x", 1.18, "Compliant"])
    _save_workbook(wb, "07_loan_schedule.xlsx")


def make_tax_computation():
    wb = Workbook()
    ws = wb.active
    ws.title = "Income_Tax"
    ws.append(["Particulars", "Amount (NPR)", "Note"])
    rows = [
        ("Profit before tax (as per accounts)", 60_300_000, ""),
        ("Add: Disallowable expenses", 2_400_000, "Donations, fines"),
        ("Add: Depreciation as per books", 18_900_000, ""),
        ("Less: Depreciation as per tax law", (21_500_000), "Accelerated rates"),
        ("Less: Tax-exempt income", (500_000), "Interest on tax-free bonds"),
        ("Taxable income", 59_600_000, ""),
        ("Income tax @ 25%", 14_900_000, ""),
        ("Add: Surcharge", 0, ""),
        ("Total income tax", 14_900_000, ""),
        ("Less: Advance tax paid", (16_500_000), ""),
        ("Refundable / (Payable)", 1_600_000, "Refundable"),
    ]
    for r in rows:
        ws.append(r)

    ws2 = wb.create_sheet("VAT")
    ws2.append(["Month", "Output VAT (NPR)", "Input VAT (NPR)", "Net Payable (NPR)"])
    months = ["Shrawan", "Bhadra", "Ashwin", "Kartik", "Mangsir", "Poush",
              "Magh", "Falgun", "Chaitra", "Baisakh", "Jestha", "Ashad"]
    random.seed(SEED)
    for m in months:
        out = random.randint(3_000_000, 6_500_000)
        inp = random.randint(2_000_000, 5_500_000)
        ws2.append([m, out, inp, out - inp])

    ws3 = wb.create_sheet("TDS")
    ws3.append(["Section", "Nature", "Amount Deducted (NPR)", "Deposited", "Status"])
    ws3.append(["88", "Rent", 720_000, 720_000, "OK"])
    ws3.append(["89", "Contract", 540_000, 540_000, "OK"])
    ws3.append(["88", "Rent — Mangsir", 60_000, 60_000, "Late deposit — penalty paid"])
    ws3.append(["88", "Rent — Magh", 60_000, 60_000, "Late deposit — penalty paid"])
    _save_workbook(wb, "08_tax_computation.xlsx")


def make_audit_checklist():
    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist"
    ws.append(["#", "Area", "Procedure", "Done by", "Reviewed by", "WP Ref", "Status"])
    items = [
        (1, "Cash & Bank", "Obtain bank confirmations", "E_NK", "M_SL", "C-100", "Done"),
        (2, "AR", "Send positive confirmation requests to top 20 customers", "E_NK", "M_SL", "D-110", "Done"),
        (3, "AR", "Test ageing >90 days for ECL adequacy", "E_NK", "M_SL", "D-115", "In progress"),
        (4, "Inventory", "Attend physical count at 3 locations", "Team", "M_SL", "E-101", "Done"),
        (5, "Inventory", "NRV testing on slow-moving SKUs", "E_NK", "M_SL", "E-110", "In progress"),
        (6, "Revenue", "Cut-off testing — last 15 GRNs / first 15 of next year", "E_NK", "M_SL", "F-130", "Done"),
        (7, "Related parties", "Inspect Board minutes; tie to RPT listing", "E_NK", "M_SL", "G-150", "Done"),
        (8, "Loan", "Re-compute DSCR; obtain bank covenant compliance certificate", "M_SL", "P_RK", "H-160", "Pending"),
        (9, "Tax", "Review provisional self-assessment; tie to GL", "M_SL", "P_RK", "I-170", "In progress"),
        (10, "ICFR", "Walk-through P2P cycle; document key controls", "E_NK", "M_SL", "J-180", "Done"),
        (11, "Going concern", "Review cash-flow forecast; sensitivity", "M_SL", "P_RK", "K-190", "Pending"),
    ]
    for r in items:
        ws.append(r)
    _save_workbook(wb, "09_audit_checklist.xlsx")


def make_rpt_listing():
    wb = Workbook()
    ws = wb.active
    ws.title = "RPT"
    ws.append(["Date", "Related Party", "Relationship", "Nature",
               "Amount (NPR)", "Board Approved", "Benchmark Available"])
    rows = [
        ("2081-05-12", "Annapurna Holdings Pvt. Ltd.", "Common director", "Purchase of raw material", 1_500_000, "Yes", "Yes"),
        ("2081-07-08", "Annapurna Holdings Pvt. Ltd.", "Common director", "Purchase of raw material", 2_800_000, "Yes", "Yes"),
        ("2081-09-22", "Annapurna Holdings Pvt. Ltd.", "Common director", "Sales of finished goods", 1_200_000, "Yes", "Yes"),
        ("2081-11-30", "Himal Family Enterprises", "Director's family", "Purchase of consumables", 600_000, "Yes", "No"),
        ("2082-01-15", "Himal Family Enterprises", "Director's family", "Purchase of consumables", 900_000, "Yes", "No"),
        ("2082-02-20", "Himal Family Enterprises", "Director's family", "Purchase of consumables", 350_000, "No", "No"),  # not approved
        ("2082-03-28", "Annapurna Holdings Pvt. Ltd.", "Common director", "Loan given (short term)", 5_000_000, "Yes", "N/A"),
    ]
    for r in rows:
        ws.append(r)
    _save_workbook(wb, "10_related_party_transactions.xlsx")


def generate_all_xlsx():
    make_trial_balance()
    make_general_ledger()
    make_fixed_asset_register()
    make_ar_aging()
    make_inventory_listing()
    make_bank_reconciliation()
    make_loan_schedule()
    make_tax_computation()
    make_audit_checklist()
    make_rpt_listing()


# ===========================================================================
# CSV files
# ===========================================================================
def _csv(name: str, df: pd.DataFrame):
    df.to_csv(CSV_DIR / name, index=False)


def make_sales_transactions():
    rng = random.Random(SEED)
    start = date(2081, 4, 1)
    rows = []
    for i in range(120):
        d = start + timedelta(days=rng.randint(0, 360))
        cust = rng.choice(CUSTOMERS)
        amt = rng.randint(50_000, 1_800_000)
        rows.append({
            "invoice_no": f"SI/2081/{1000+i:04d}",
            "date": d.isoformat(),
            "customer_code": cust[0],
            "customer_name": cust[1],
            "amount": amt,
            "vat": round(amt * 0.13),
            "total": amt + round(amt * 0.13),
            "payment_status": rng.choices(["Paid", "Outstanding", "Late"], weights=[6, 3, 1])[0],
        })
    # Round-number unusual sale
    rows.append({"invoice_no": "SI/2081/9999", "date": "2082-03-31", "customer_code": "C004",
                 "customer_name": "Annapurna Holdings Pvt. Ltd.", "amount": 1_000_000,
                 "vat": 130_000, "total": 1_130_000, "payment_status": "Outstanding"})
    df = pd.DataFrame(rows)
    _csv("01_sales_transactions.csv", df)


def make_purchase_transactions():
    rng = random.Random(SEED + 1)
    start = date(2081, 4, 1)
    rows = []
    for i in range(100):
        d = start + timedelta(days=rng.randint(0, 360))
        v = rng.choice(VENDORS)
        amt = rng.randint(40_000, 1_500_000)
        rows.append({
            "invoice_no": f"PI/2081/{2000+i:04d}",
            "date": d.isoformat(),
            "vendor_code": v[0],
            "vendor_name": v[1],
            "amount": amt,
            "vat": round(amt * 0.13),
            "approval_status": rng.choices(["Approved", "Pending", "Sole-source"], weights=[8, 1, 1])[0],
        })
    # Inject duplicate invoice numbers (teaching trigger)
    dup = rows[10].copy()
    dup["vendor_code"] = "V011"
    dup["vendor_name"] = "Kathmandu Steel Supplier P. Ltd."   # spelling variant
    rows.append(dup)
    dup2 = rows[25].copy()
    rows.append(dup2)
    # A few related-party round-number purchases
    rows.append({"invoice_no": "PI/2081/9101", "date": "2082-02-28", "vendor_code": "V004",
                 "vendor_name": "Annapurna Holdings Pvt. Ltd.", "amount": 2_000_000,
                 "vat": 260_000, "approval_status": "Approved"})
    rows.append({"invoice_no": "PI/2081/9102", "date": "2082-03-15", "vendor_code": "V010",
                 "vendor_name": "Himal Family Enterprises", "amount": 500_000,
                 "vat": 65_000, "approval_status": "Sole-source"})
    df = pd.DataFrame(rows)
    _csv("02_purchase_transactions.csv", df)


def make_journal_entries():
    rng = random.Random(SEED + 2)
    start = date(2081, 4, 1)
    rows = []
    for i in range(80):
        d = start + timedelta(days=rng.randint(0, 360))
        amt = rng.randint(10_000, 500_000)
        rows.append({
            "je_no": f"JE-{1000+i:04d}",
            "date": d.isoformat(),
            "account_debit": rng.choice([a[0] for a in ACCOUNTS]),
            "account_credit": rng.choice([a[0] for a in ACCOUNTS]),
            "amount": amt,
            "narration": "Routine posting",
            "posted_by": rng.choice([e[0] for e in EMPLOYEES]),
            "approved_by": rng.choice([e[0] for e in EMPLOYEES]),
        })
    # Unusual entries
    rows.append({"je_no": "JE-9001", "date": "2082-03-31", "account_debit": "5700",
                 "account_credit": "1100", "amount": 1_000_000,
                 "narration": "Adjustment", "posted_by": "E008", "approved_by": "E008"})  # self-approved + round
    rows.append({"je_no": "JE-9002", "date": "2082-03-31", "account_debit": "4100",
                 "account_credit": "1200", "amount": 250_000,
                 "narration": "Reversal", "posted_by": "E004", "approved_by": ""})       # no approver
    df = pd.DataFrame(rows)
    _csv("03_journal_entries.csv", df)


def make_vendor_master():
    rows = []
    for v in VENDORS:
        rows.append({
            "vendor_code": v[0],
            "vendor_name": v[1],
            "pan": v[2] if v[2] else "",
            "related_party": "Yes" if v[3] else "No",
            "active": "Yes",
        })
    df = pd.DataFrame(rows)
    _csv("04_vendor_master.csv", df)


def make_customer_master():
    rows = []
    for c in CUSTOMERS:
        rows.append({
            "customer_code": c[0],
            "customer_name": c[1],
            "pan": c[2] if c[2] else "",
            "credit_limit": random.choice([2_000_000, 5_000_000, 10_000_000]),
            "active": "Yes",
        })
    df = pd.DataFrame(rows)
    _csv("05_customer_master.csv", df)


def make_payroll_summary():
    rng = random.Random(SEED + 3)
    rows = []
    for e in EMPLOYEES:
        base = rng.randint(40_000, 200_000)
        ssf = round(base * 0.11)
        tds = round(base * 0.10)
        net = base - ssf - tds
        rows.append({
            "employee_code": e[0],
            "name": e[1],
            "designation": e[2],
            "monthly_gross": base,
            "ssf_deduction": ssf,
            "tds": tds,
            "net_pay": net,
        })
    df = pd.DataFrame(rows)
    _csv("06_payroll_summary.csv", df)


def make_expense_claims():
    rng = random.Random(SEED + 4)
    start = date(2081, 4, 1)
    rows = []
    for i in range(40):
        d = start + timedelta(days=rng.randint(0, 360))
        emp = rng.choice(EMPLOYEES)
        amt = rng.randint(2_000, 80_000)
        rows.append({
            "claim_no": f"EXP-{500+i:04d}",
            "date": d.isoformat(),
            "employee_code": emp[0],
            "employee_name": emp[1],
            "category": rng.choice(["Travel", "Meals", "Local Transport", "Office Supplies", "Client Entertainment"]),
            "amount": amt,
            "receipt_attached": rng.choices(["Yes", "No"], weights=[9, 1])[0],
            "approved": rng.choices(["Yes", "No"], weights=[19, 1])[0],
        })
    df = pd.DataFrame(rows)
    _csv("07_expense_claims.csv", df)


def make_budget_vs_actual():
    rows = [
        ("Revenue", 450_000_000, 486_234_500),
        ("COGS", 340_000_000, 364_676_000),
        ("Salaries", 36_000_000, 38_500_000),
        ("Rent", 7_200_000, 7_200_000),
        ("Utilities", 4_000_000, 4_600_000),
        ("Depreciation", 19_000_000, 18_900_000),
        ("Interest", 18_500_000, 19_110_000),
        ("Other Operating", 9_000_000, 10_758_500),
        ("Tax", 13_500_000, 14_900_000),
    ]
    df = pd.DataFrame(rows, columns=["line_item", "budget", "actual"])
    df["variance"] = df["actual"] - df["budget"]
    df["variance_pct"] = (df["variance"] / df["budget"] * 100).round(2)
    _csv("08_budget_vs_actual.csv", df)


def make_inventory_movements():
    rng = random.Random(SEED + 5)
    start = date(2081, 4, 1)
    rows = []
    skus = ["RM-001", "RM-014", "RM-027", "WIP-101", "FG-302", "FG-118", "FG-205", "CONS-09", "FG-509"]
    for i in range(150):
        d = start + timedelta(days=rng.randint(0, 360))
        rows.append({
            "date": d.isoformat(),
            "sku": rng.choice(skus),
            "movement_type": rng.choice(["IN", "OUT", "TRANSFER", "ADJUSTMENT"]),
            "qty": rng.randint(1, 200),
            "ref": f"MV-{3000+i:05d}",
        })
    df = pd.DataFrame(rows)
    _csv("09_inventory_movements.csv", df)


def make_risk_register():
    rows = [
        ("R001", "Revenue cut-off error", "Financial reporting", "High", "Medium", "Cut-off testing at YE"),
        ("R002", "Related-party pricing not at arm's length", "Tax / Compliance", "Medium", "High", "Benchmark documentation"),
        ("R003", "Loan covenant breach", "Treasury", "Medium", "High", "Quarterly DSCR monitoring"),
        ("R004", "Inventory obsolescence under-provisioning", "Financial reporting", "Medium", "Medium", "NRV testing"),
        ("R005", "Duplicate vendor master entries", "Procurement", "Low", "Medium", "Master data clean-up"),
        ("R006", "Late TDS deposit", "Tax", "Low", "Medium", "Automated reminder"),
        ("R007", "Sole-source procurement above threshold", "Procurement", "Medium", "Medium", "Tender process enforcement"),
        ("R008", "Unauthorised journal entries near YE", "ICFR", "Medium", "High", "Maker-checker on JE"),
        ("R009", "Foreign exchange exposure", "Treasury", "Low", "Medium", "Forward cover policy"),
        ("R010", "Going concern under stress scenarios", "Strategic", "Low", "High", "Sensitivity in cash forecast"),
    ]
    df = pd.DataFrame(rows, columns=[
        "risk_id", "risk_description", "category", "likelihood", "impact", "mitigation"
    ])
    _csv("10_risk_register.csv", df)


def generate_all_csv():
    make_sales_transactions()
    make_purchase_transactions()
    make_journal_entries()
    make_vendor_master()
    make_customer_master()
    make_payroll_summary()
    make_expense_claims()
    make_budget_vs_actual()
    make_inventory_movements()
    make_risk_register()


# ===========================================================================
# Data dictionary
# ===========================================================================
DATA_DICTIONARY = """\
# Data Dictionary — ICAN CA RAG Training (Synthetic Data)

All data below is **synthetic**. The fictional company is
**Himal Trading & Manufacturing Pvt. Ltd.** (PAN 300123456), with year-end
**2082-03-31** (Nepali FY 2081/82). No real entity is referenced.

> **Teaching-friendly issues** have been intentionally seeded across the data set.
> Look for them with your participants.

---

## PDF documents — `data/generated/pdf/`

| # | File | Purpose | Notable items to discover |
|---|---|---|---|
| 01 | `01_annual_report_extract.pdf`        | Directors' report, KPIs, borrowings, RPT, going concern | Cross-refs to loan agreement and RPT listing |
| 02 | `02_audit_planning_memo.pdf`          | Risks, materiality, team                                | 5 significant risks; performance materiality NPR 15.75 lakh |
| 03 | `03_internal_control_policy.pdf`      | Segregation, approval matrix, vendor master controls    | Three-tier approval thresholds; exception register |
| 04 | `04_procurement_policy.pdf`           | 3-quote rule, 3-way match, related-party rule           | Tender required > NPR 10,00,000 |
| 05 | `05_tax_compliance_memo.pdf`          | IT, VAT, TDS status                                     | Open exposures incl. transfer pricing |
| 06 | `06_loan_agreement_summary.pdf`       | Covenants, security, EoD                                | DSCR ≥ 1.25x, D/E ≤ 2.0x, CR ≥ 1.10x |
| 07 | `07_board_minutes.pdf`                | RPT approvals, IA findings                              | Family entities approved; sole-source flagged |
| 08 | `08_management_letter.pdf`            | Findings from audit                                     | Duplicate vendors, missing PAN, IT shared IDs |
| 09 | `09_nfrs_accounting_policy_note.pdf`  | Revenue/inventory/PPE/FI/RPT policies                   | NFRS 15, NAS 2, NAS 16, NFRS 9, NAS 24 |
| 10 | `10_inventory_observation_memo.pdf`   | Physical count exceptions                               | Variances at 3 locations; slow-moving 4.7% |

## Excel workbooks — `data/generated/xlsx/`

| # | File | Sheets | Notes |
|---|---|---|---|
| 01 | `01_trial_balance.xlsx`              | TB                              | 20 accounts, totals row |
| 02 | `02_general_ledger_sample.xlsx`      | GL_Sales / GL_Purchases / GL_Adjustments | JE-0099 round-number, late-night, junior-posted |
| 03 | `03_fixed_asset_register.xlsx`       | FAR                             | 8 assets with NBV |
| 04 | `04_accounts_receivable_aging.xlsx`  | AR_Aging                        | C004 flat aging (RPT signal); C008 fully overdue |
| 05 | `05_inventory_listing.xlsx`          | Inventory                       | RM-027 / FG-205 / FG-509 slow / obsolete |
| 06 | `06_bank_reconciliation.xlsx`        | Bank_Recon / Outstanding_Items  | Stale cheques outstanding > 170 days |
| 07 | `07_loan_schedule.xlsx`              | Loan_Schedule / Covenants       | EMI table + covenant tracker |
| 08 | `08_tax_computation.xlsx`            | Income_Tax / VAT / TDS          | Income-tax refund of NPR 16 lakh; late TDS |
| 09 | `09_audit_checklist.xlsx`            | Checklist                       | 11 audit procedures with status |
| 10 | `10_related_party_transactions.xlsx` | RPT                             | One un-approved entry on 2082-02-20 |

## CSV transactional files — `data/generated/csv/`

| # | File | Rows ≈ | Teaching trigger |
|---|---|---|---|
| 01 | `01_sales_transactions.csv`     | 121 | Round-number RPT sale on year-end |
| 02 | `02_purchase_transactions.csv`  | 103 | **Duplicate invoice numbers**, vendor-name variants |
| 03 | `03_journal_entries.csv`        |  82 | Self-approved JE-9001; missing approver JE-9002 |
| 04 | `04_vendor_master.csv`          |  12 | V006 missing PAN; V011/V012 are duplicates of V001/V003 |
| 05 | `05_customer_master.csv`        |   8 | C006 missing PAN; C004 is a related party |
| 06 | `06_payroll_summary.csv`        |   8 | Salary, SSF, TDS, net |
| 07 | `07_expense_claims.csv`         |  40 | Some without receipts; rare un-approved claims |
| 08 | `08_budget_vs_actual.csv`       |   9 | Revenue +8%, COGS +7%, Utilities +15% |
| 09 | `09_inventory_movements.csv`    | 150 | Stock movement journal |
| 10 | `10_risk_register.csv`          |  10 | Used in capstone risk-mapping |

---

## How to use this in the training

* **Notebook 03** — open every file type and preview.
* **Notebook 05** — RAG over PDFs only.
* **Notebook 07** — multi-document RAG combining PDFs + structured ledgers.
* **Notebook 09** — pick three failures and reproduce them (e.g. ask a question outside the corpus).
* **Notebook 10/11** — build a graph: Vendor → Invoice → Approver; query related-party paths.
* **Notebook 12/13** — capstone agent that pulls from PDFs, queries CSVs, and walks the graph.

---

## Reproducibility

All data is deterministically generated from seed `20260526` in
`src/data_generation.py`. Re-running the script overwrites the existing files.
"""


def write_data_dictionary():
    path = DATA_DIR / "DATA_DICTIONARY.md"
    path.write_text(DATA_DICTIONARY, encoding="utf-8")


# ===========================================================================
# Main
# ===========================================================================
def main():
    print("Generating ICAN CA RAG training dataset...")
    generate_all_pdfs()
    print(f"[ok] PDFs written to {PDF_DIR}/ ({len(list(PDF_DIR.glob('*.pdf')))} files)")
    generate_all_xlsx()
    print(f"[ok] XLSX written to {XLSX_DIR}/ ({len(list(XLSX_DIR.glob('*.xlsx')))} files)")
    generate_all_csv()
    print(f"[ok] CSVs written to {CSV_DIR}/ ({len(list(CSV_DIR.glob('*.csv')))} files)")
    write_data_dictionary()
    print(f"[ok] Data dictionary written to {DATA_DIR / 'DATA_DICTIONARY.md'}")
    print("All synthetic data generated successfully.")


if __name__ == "__main__":
    main()
